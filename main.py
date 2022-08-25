import pyautogui
from openpyxl import load_workbook
from selenium import webdriver
from webdrivermanager import ChromeDriverManager
import pygetwindow as gw
import os
import glob
import win32printing
import win32com.client
import win32print
import time
from openpyxl.styles.borders import Border, Side
#dia chi lich su hanh trinh cua dinh vi
url = 'http://api.eup.net.vn:8000/ctynguyenngoc/history'
#dia chi file chi tiet (ghi tat add1)
add1 = "C:/Users/Work/Desktop\chitietlotrinh.xlsx"
#load excel file
workbook = load_workbook(filename=add1)
#open workbook
#tensheet = workbook.sheetnames
#print(tensheet)
#goi ten sheet
chitiet = workbook["Sheet1"]
datasheet = workbook["Data"]
def lammoi():
    chitiet["B5"] = ""
    chitiet["B7"] = ""
    chitiet.delete_rows(idx=9, amount=1000)
    workbook.save(add1)
#dem so dong o sheet data
mr = datasheet.max_row
#bat dau tung dong cua file data
for i in range(2,mr+1):
    lammoi()
    soxe = datasheet.cell(i,2).value
    shipment = datasheet.cell(i,1).value
    chitiet["B5"] = soxe
    chitiet["B7"] = shipment
    workbook.save(add1)
    a = gw.getWindowsWithTitle('19136 - CÔNG TY CỔ PHẦN NGUYỄN NGỌC LOGISTICS - v1.0.15.217 - Google Chrome')[0]
    a.activate()
    pyautogui.moveTo(1234,202)
    pyautogui.leftClick()
    pyautogui.write(soxe)
    pyautogui.moveTo(1186, 264, duration=1)
    pyautogui.leftClick()
    pyautogui.moveTo(1114, 354, duration =1)
    pyautogui.leftClick()
    pyautogui.moveTo(1343, 438, duration =2)
    pyautogui.leftClick()
    pyautogui.moveTo(1250, 260, duration= 3)
    pyautogui.scroll(-300)
    pyautogui.moveTo(1121, 367, duration=2)
    pyautogui.leftClick()
    pyautogui.moveTo(500, 400, duration=2)
    pyautogui.leftClick()
    list_of_files = glob.glob("C:/Users/Work/Downloads/*xlsx")  # * means all if need specific format then *.csv
    add2 = max(list_of_files, key=os.path.getctime)
    print(add2)
    # load excel file 2
    wb2 = load_workbook(filename=add2)
    r2 = wb2.sheetnames
    print(r2)
    r = wb2.active
    for i in range(4,250):
        chitiet.cell(5+i,1).value=r.cell(i,1).value
        chitiet.cell(5 + i,2).value= r.cell(i, 2).value
        chitiet.cell(5 + i, 3).value = r.cell(i, 3).value
        chitiet.cell(5+i, 4).value= r.cell(i, 4).value
        chitiet.cell(5+i, 5).value = r.cell(i, 5).value
    #o = win32com.client.Dispatch('Excel.Application')
    #o.visible = True
    #wb = o.Workbooks.Open(add1)
    #ws = wb.Worksheets(["Sheet1"])
    #ws.printout()
    #o.Quit()
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for dong in range (9,300):
        for cot in range(1,6):
            chitiet.cell(dong,cot).border=thin_border
    workbook.save(add1)
    os.startfile(add1, 'print')
